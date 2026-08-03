"""Build the canonical, auditable FRF dataset from the original workbook.

The original Excel workbook is treated as read-only input.  This module keeps
source sheet/row/column coordinates, recomputes derived dB/degree values from
linear magnitude and radians, and writes deterministic artifacts for later
model identification and RL environment work.
"""

from __future__ import annotations

import hashlib
import json
import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


CURRENT_ROWS = range(21, 39)
SPEED_BLOCK_START_COLUMNS = (1, 8, 15, 22, 29, 36)
SPEED_AMPLITUDES = (10, 50, 100, 250, 500, 700)
SPEED_ROLES = {
    10: "robustness_ood",
    50: "robustness_suspicious",
    100: "train",
    250: "train",
    500: "train",
    700: "validation",
}
TASK_ID = "cgs_turntable_001"
FRF_FEATURE_NAMES = (
    "log_frequency",
    "magnitude_db",
    "sin_phase",
    "cos_phase",
    "quality_mask",
)


def _as_float(value: Any) -> float:
    result = float(value)
    if not math.isfinite(result):
        raise ValueError(f"non-finite source value: {value!r}")
    return result


def _phase_deg_clean(phase_rad: np.ndarray) -> np.ndarray:
    """Return phase in a stable [-180, 180) branch after frequency sorting."""

    unwrapped = np.unwrap(phase_rad)
    degrees = np.rad2deg(unwrapped)
    return (degrees + 180.0) % 360.0 - 180.0


def _records_to_frame(records: list[dict[str, Any]]) -> pd.DataFrame:
    frame = pd.DataFrame.from_records(records)
    if frame.empty:
        raise ValueError("no FRF records were parsed")

    frame["magnitude_db"] = 20.0 * np.log10(frame["magnitude_linear"].astype(float))
    frame = frame.sort_values(
        ["loop", "amplitude_mA", "frequency_hz", "source_row"],
        kind="mergesort",
        na_position="last",
    ).reset_index(drop=True)
    phase = frame["phase_rad_raw"].astype(float).to_numpy()
    frame["phase_deg_raw"] = np.rad2deg(phase)
    cleaned_phase = frame["phase_deg_raw"].to_numpy(dtype=float).copy()
    for _, group in frame[frame["loop"] == "speed"].groupby("amplitude_mA", sort=False):
        indices = group.index.to_numpy()
        cleaned_phase[indices] = _phase_deg_clean(phase[indices])
    frame["phase_deg_clean"] = cleaned_phase
    frame["phase_rad_clean"] = np.deg2rad(frame["phase_deg_clean"])
    frame["point_index"] = frame.groupby(
        ["loop", "amplitude_mA"], sort=False
    ).cumcount()

    group_columns = ["loop", "amplitude_mA"]
    frame["frequency_strictly_increasing"] = frame.groupby(group_columns)[
        "frequency_hz"
    ].transform(lambda values: values.diff().iloc[1:].gt(0).all() if len(values) > 1 else True)
    frame["duplicate_frequency"] = frame.duplicated(group_columns + ["frequency_hz"], keep=False)
    frame["finite_values"] = np.isfinite(
        frame[["frequency_hz", "magnitude_linear", "phase_rad_raw", "magnitude_db"]]
    ).all(axis=1)
    frame["quality_ok"] = frame["finite_values"] & ~frame["duplicate_frequency"]

    def quality_flags(row: pd.Series) -> str:
        flags: list[str] = []
        if not bool(row["finite_values"]):
            flags.append("nonfinite")
        if bool(row["duplicate_frequency"]):
            flags.append("duplicate_frequency")
        if not bool(row["frequency_strictly_increasing"]):
            flags.append("frequency_order_issue")
        if row["loop"] == "speed" and row["amplitude_mA"] == 10:
            flags.append("ood_amplitude")
        if row["loop"] == "speed" and row["amplitude_mA"] == 50 and row["point_index"] == 0:
            flags.append("suspicious_first_point")
        return ";".join(flags)

    frame["quality_flags"] = frame.apply(quality_flags, axis=1)
    return frame


def parse_original_workbook(path: Path) -> pd.DataFrame:
    """Parse the three known sheets by index, avoiding locale-dependent names."""

    workbook = load_workbook(path, data_only=False, read_only=True)
    if len(workbook.worksheets) < 3:
        raise ValueError(f"expected three FRF worksheets, got {len(workbook.worksheets)}")

    current_sheet = workbook.worksheets[0]
    speed_sheet = workbook.worksheets[1]
    position_sheet = workbook.worksheets[2]
    records: list[dict[str, Any]] = []

    for row in CURRENT_ROWS:
        frequency = current_sheet.cell(row, 7).value
        magnitude = current_sheet.cell(row, 8).value
        phase = current_sheet.cell(row, 9).value
        if frequency is None:
            continue
        records.append(
            {
                "loop": "current",
                "amplitude_mA": np.nan,
                "excitation": "500mV",
                "role": "reference",
                "frequency_hz": _as_float(frequency),
                "magnitude_linear": _as_float(magnitude),
                "phase_rad_raw": _as_float(phase),
                "source_sheet": current_sheet.title,
                "source_row": row,
                "source_columns": "G:I",
            }
        )

    for start_column, amplitude in zip(SPEED_BLOCK_START_COLUMNS, SPEED_AMPLITUDES):
        for row in range(3, 18):
            frequency = speed_sheet.cell(row, start_column).value
            magnitude = speed_sheet.cell(row, start_column + 1).value
            phase = speed_sheet.cell(row, start_column + 2).value
            if frequency is None:
                continue
            records.append(
                {
                    "loop": "speed",
                    "amplitude_mA": float(amplitude),
                    "excitation": f"{amplitude}mA",
                    "role": SPEED_ROLES[amplitude],
                    "frequency_hz": _as_float(frequency),
                    "magnitude_linear": _as_float(magnitude),
                    "phase_rad_raw": _as_float(phase),
                    "source_sheet": speed_sheet.title,
                    "source_row": row,
                    "source_columns": f"{start_column}:{start_column + 2}",
                }
            )

    for row in range(1, 16):
        frequency = position_sheet.cell(row, 6).value
        magnitude = position_sheet.cell(row, 7).value
        phase = position_sheet.cell(row, 8).value
        records.append(
            {
                "loop": "position",
                "amplitude_mA": np.nan,
                "excitation": "position",
                "role": "reference",
                "frequency_hz": _as_float(frequency),
                "magnitude_linear": _as_float(magnitude),
                "phase_rad_raw": _as_float(phase),
                "source_sheet": position_sheet.title,
                "source_row": row,
                "source_columns": "F:H",
            }
        )

    frame = _records_to_frame(records)
    frame["amplitude_mA"] = frame["amplitude_mA"].astype("Float64")
    return frame


def _interleaved_response(frame: pd.DataFrame) -> np.ndarray:
    return frame[["magnitude_db", "phase_deg_clean"]].to_numpy(dtype=float).reshape(-1)


def _canonical_arrays(frame: pd.DataFrame) -> dict[str, np.ndarray]:
    current = frame[frame["loop"] == "current"].sort_values("frequency_hz")
    position = frame[frame["loop"] == "position"].sort_values("frequency_hz")
    speed_parts = []
    speed_frequencies = []
    amplitudes = []
    for amplitude in SPEED_AMPLITUDES:
        part = frame[(frame["loop"] == "speed") & (frame["amplitude_mA"] == amplitude)].sort_values(
            "frequency_hz"
        )
        amplitudes.append(float(amplitude))
        speed_parts.append(part[["magnitude_db", "phase_deg_clean"]].to_numpy(dtype=float))
        speed_frequencies.append(part["frequency_hz"].to_numpy(dtype=float))

    current_response = current[["magnitude_db", "phase_deg_clean"]].to_numpy(dtype=float)
    position_response = position[["magnitude_db", "phase_deg_clean"]].to_numpy(dtype=float)
    speed_response = np.stack(speed_parts)
    input_vector = np.concatenate(
        [
            current_response.reshape(-1),
            speed_response[4].reshape(-1),
            position_response.reshape(-1),
        ]
    )
    scale = np.tile(np.array([40.0, 180.0]), input_vector.size // 2)
    return {
        "current_frequency_hz": current["frequency_hz"].to_numpy(dtype=float),
        "current_response_db_deg": current_response,
        "speed_amplitudes_mA": np.asarray(amplitudes, dtype=float),
        "speed_frequency_hz": np.stack(speed_frequencies),
        "speed_response_db_deg": speed_response,
        "position_frequency_hz": position["frequency_hz"].to_numpy(dtype=float),
        "position_response_db_deg": position_response,
        "input_vector": input_vector,
        "input_vector_scaled": input_vector / scale,
    }


def _task_feature_matrix(frame: pd.DataFrame) -> np.ndarray:
    ordered = frame.sort_values("frequency_hz", kind="mergesort")
    phase = ordered["phase_rad_clean"].to_numpy(dtype=float)
    suspicious = ordered["quality_flags"].str.contains(
        "suspicious_first_point", regex=False, na=False
    )
    quality_mask = (
        ordered["quality_ok"].to_numpy(dtype=bool) & ~suspicious.to_numpy(dtype=bool)
    ).astype(np.float64)
    return np.column_stack(
        [
            np.log10(ordered["frequency_hz"].to_numpy(dtype=float)),
            ordered["magnitude_db"].to_numpy(dtype=float),
            np.sin(phase),
            np.cos(phase),
            quality_mask,
        ]
    )


def _frf_task_arrays(frame: pd.DataFrame) -> dict[str, np.ndarray]:
    current = _task_feature_matrix(frame[frame["loop"] == "current"])
    position = _task_feature_matrix(frame[frame["loop"] == "position"])

    speed_parts: list[np.ndarray] = []
    speed_suspicious_masks: list[np.ndarray] = []
    maximum_amplitude = float(max(SPEED_AMPLITUDES))
    for amplitude in SPEED_AMPLITUDES:
        part = frame[
            (frame["loop"] == "speed") & (frame["amplitude_mA"] == amplitude)
        ].sort_values("frequency_hz", kind="mergesort")
        features = _task_feature_matrix(part)
        amplitude_column = np.full((len(features), 1), amplitude / maximum_amplitude)
        speed_parts.append(np.column_stack([features, amplitude_column]))
        speed_suspicious_masks.append(
            part["quality_flags"]
            .str.contains("suspicious_first_point", regex=False, na=False)
            .to_numpy(dtype=np.uint8)
        )

    speed = np.stack(speed_parts)
    context_vector = np.concatenate([current.reshape(-1), speed.reshape(-1), position.reshape(-1)])
    return {
        "schema_version": np.asarray(1, dtype=np.int16),
        "task_id": np.asarray(TASK_ID),
        "loop_names": np.asarray(["current", "speed", "position"]),
        "base_feature_names": np.asarray(FRF_FEATURE_NAMES),
        "speed_feature_names": np.asarray((*FRF_FEATURE_NAMES, "amplitude_norm")),
        "current_frf": current,
        "speed_frf": speed,
        "position_frf": position,
        "speed_amplitudes_mA": np.asarray(SPEED_AMPLITUDES, dtype=float),
        "speed_roles": np.asarray([SPEED_ROLES[value] for value in SPEED_AMPLITUDES]),
        "speed_suspicious_mask": np.stack(speed_suspicious_masks),
        "context_vector": context_vector,
    }


def _json_value(value: Any) -> Any:
    if isinstance(value, (np.integer, np.floating)):
        value = value.item()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, np.ndarray):
        return [_json_value(item) for item in value.tolist()]
    return value


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_figures(frame: pd.DataFrame, figure_dir: Path) -> list[str]:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    figure_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[str] = []
    for loop, filename, title in (
        ("current", "current_frf.png", "Current-loop FRF"),
        ("position", "position_frf.png", "Position-loop FRF"),
    ):
        part = frame[frame["loop"] == loop].sort_values("frequency_hz")
        fig, axes = plt.subplots(2, 1, figsize=(8, 6), sharex=True)
        axes[0].semilogx(part["frequency_hz"], part["magnitude_db"], "o-")
        axes[0].set_ylabel("Magnitude (dB)")
        axes[1].semilogx(part["frequency_hz"], part["phase_deg_clean"], "o-")
        axes[1].set_ylabel("Phase (deg)")
        axes[1].set_xlabel("Frequency (Hz)")
        fig.suptitle(title)
        fig.tight_layout()
        output = figure_dir / filename
        fig.savefig(output, dpi=150)
        plt.close(fig)
        outputs.append(str(output))

    fig, axes = plt.subplots(2, 1, figsize=(8, 7), sharex=True)
    for amplitude in SPEED_AMPLITUDES:
        part = frame[(frame["loop"] == "speed") & (frame["amplitude_mA"] == amplitude)].sort_values(
            "frequency_hz"
        )
        axes[0].semilogx(part["frequency_hz"], part["magnitude_db"], "o-", label=f"{amplitude} mA")
        axes[1].semilogx(part["frequency_hz"], part["phase_deg_clean"], "o-", label=f"{amplitude} mA")
    axes[0].set_ylabel("Magnitude (dB)")
    axes[1].set_ylabel("Phase (deg)")
    axes[1].set_xlabel("Frequency (Hz)")
    axes[0].legend(ncol=3, fontsize=8)
    fig.suptitle("Speed-loop FRF by excitation amplitude")
    fig.tight_layout()
    output = figure_dir / "speed_frf_by_amplitude.png"
    fig.savefig(output, dpi=150)
    plt.close(fig)
    outputs.append(str(output))
    return outputs


def build_processed_data(project_root: Path) -> dict[str, Any]:
    data_dir = project_root / "data"
    original_path = data_dir / "original" / "OSN600 9# CGS转台开环频响.xlsx"
    seed_path = data_dir / "CGS转台_RL训练种子数据.npz"
    output_dir = data_dir / "processed"
    figure_dir = output_dir / "figures"
    output_dir.mkdir(parents=True, exist_ok=True)

    frame = parse_original_workbook(original_path)
    canonical = _canonical_arrays(frame)
    seed = np.load(seed_path, allow_pickle=False)

    reference_comparison = {}
    for key, value in canonical.items():
        if key not in seed.files:
            continue
        reference_comparison[key] = {
            "shape": list(value.shape),
            "max_abs_difference": float(np.max(np.abs(value - seed[key]))),
        }

    frame.to_parquet(output_dir / "frf_long.parquet", index=False)
    np.savez_compressed(output_dir / "frf_canonical.npz", **canonical)

    task_arrays = _frf_task_arrays(frame)
    task_path = output_dir / "frf_tasks.npz"
    np.savez_compressed(task_path, **task_arrays)
    task_manifest = {
        "schema_version": 1,
        "task_id": TASK_ID,
        "description": (
            "Offline measured three-loop FRF artifact for model-mismatch "
            "diagnostics; excluded from SAC training."
        ),
        "data_role": "offline_model_mismatch_diagnostic_only",
        "training_policy": {
            "included_in_observation": False,
            "included_in_reward": False,
            "included_in_replay_buffer": False,
            "included_in_candidate_selection": False,
            "included_in_server_packages": False,
        },
        "arrays": {
            key: {"shape": list(value.shape), "dtype": str(value.dtype)}
            for key, value in task_arrays.items()
        },
        "feature_schema": {
            "base": list(FRF_FEATURE_NAMES),
            "speed": [*FRF_FEATURE_NAMES, "amplitude_norm"],
            "amplitude_norm": "amplitude_mA / 700",
            "phase": "sin/cos of cleaned phase in radians",
        },
        "context_vector": {
            "shape": [705],
            "order": ["current_frf", "speed_frf", "position_frf"],
            "slices": {
                "current_frf": [0, 90],
                "speed_frf": [90, 630],
                "position_frf": [630, 705],
            },
        },
        "speed_roles": {str(key): value for key, value in SPEED_ROLES.items()},
        "quality_policy": {
            "retained_points": 123,
            "quality_mask_zero": 1,
            "rule": "The suspicious first 50 mA point is retained with quality_mask=0.",
        },
        "provenance": {
            "original_workbook_sha256": _sha256(original_path),
            "frf_canonical_sha256": _sha256(output_dir / "frf_canonical.npz"),
            "frf_long_sha256": _sha256(output_dir / "frf_long.parquet"),
            "frf_tasks_sha256": _sha256(task_path),
        },
        "limitations": [
            (
                "The retained 705-element context_vector is a legacy/offline "
                "representation and is not loaded by PIDTuningEnv."
            ),
            "This file contains measured diagnostics, not RL transitions or optimal controller labels.",
            "Position FRF is retained as an independent measured channel; no speed/s topology is assumed.",
            "Speed amplitude roles are provisional experiment roles and remain explicit in the task.",
        ],
    }
    _write_json(output_dir / "frf_tasks_manifest.json", task_manifest)


    baseline_values = seed["baseline_analog_values"]
    _write_json(
        output_dir / "controller_baselines.json",
        {
            "source": "CGS转台_RL训练种子数据.npz",
            "output_names": seed["output_names"].tolist(),
            "values": [_json_value(item) for item in baseline_values],
            "note": "Missing controller values remain null; no PID/DOBC values are fabricated.",
        },
    )

    quality_columns = [
        "loop",
        "amplitude_mA",
        "point_index",
        "frequency_hz",
        "source_sheet",
        "source_row",
        "source_columns",
        "role",
        "finite_values",
        "frequency_strictly_increasing",
        "duplicate_frequency",
        "quality_ok",
        "quality_flags",
    ]
    frame[quality_columns].to_csv(output_dir / "quality_flags.csv", index=False, encoding="utf-8-sig")
    figures = _write_figures(frame, figure_dir)

    manifest = {
        "schema_version": 1,
        "source_files": {
            "original_workbook": {
                "path": str(original_path),
                "sha256": _sha256(original_path),
            },
            "seed_npz": {"path": str(seed_path), "sha256": _sha256(seed_path)},
        },
        "counts": {
            "total_points": int(len(frame)),
            "current_points": int((frame["loop"] == "current").sum()),
            "speed_points": int((frame["loop"] == "speed").sum()),
            "position_points": int((frame["loop"] == "position").sum()),
            "quality_ok_points": int(frame["quality_ok"].sum()),
        },
        "speed_roles": {str(key): value for key, value in SPEED_ROLES.items()},
        "processing_rules": [
            "Original workbook is read-only; derived values are recomputed from linear magnitude and radians.",
            "Current-loop canonical data uses original rows 21-38 (After Fix data) and is sorted by frequency.",
            "Speed-loop blocks are sorted by frequency and use a [-180,180) clean phase; current/position phase retains the source branch.",
            "The 10 mA speed block is OOD and the first 50 mA point is retained with a suspicious quality flag.",
            "No reward, transition, PID, DOBC, or missing controller labels are fabricated.",
        ],
        "reference_comparison_to_seed_npz": reference_comparison,
        "artifacts": [
            "frf_long.parquet",
            "frf_canonical.npz",
            "frf_tasks.npz",
            "frf_tasks_manifest.json",
            "controller_baselines.json",
            "quality_flags.csv",
            *[str(Path(path).relative_to(output_dir)) for path in figures],
        ],
    }
    _write_json(output_dir / "manifest.json", manifest)
    return manifest


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[2]
    result = build_processed_data(root)
    print(json.dumps(result["counts"], ensure_ascii=False))
